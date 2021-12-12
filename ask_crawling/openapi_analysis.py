from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

# Company that you find
file = '00126380_annual.xlsx'

wb = load_workbook(file)
item_list = []


# 연도별 Data를 가져올 함수 설정
def get_data(r, ws, list):
    c = 2
    while ws.cell(2, c).value != None:
        y = ws.cell(1, c).value
        if y != None and y[:1] == '2':
            for i in range(10):
                list.append(ws.cell(r, c + i).value)
            break
        c += 1


# 제무상태표, 손익계산서, 현금흐름을 가져올 수 있는 함수설정
def get_fs_data(ws, fs_items, df_list):
    for fs_item in fs_items:
        temp_list = []
        r = 4
        while ws.cell(r, 2).value != None:
            if ws.cell(r, 2).value == fs_item:
                key = ws.cell(r, 2).value
                item = ws.cell(r, 3).value
                item_list.append(item)
                get_data(r, ws, temp_list)
            r += 1
        df = pd.DataFrame(temp_list, columns=[key], index=y_list)
        df_list.append(df)


# 재무상태표를 통해 년도를 List로 만들기
y_list = []
get_data(1, wb['Data_bs'], y_list)

# 각 데이터가 담긴 Dataframe 리스트를 담을 변수 지정
df_list = []

# 재무상태표 Data 가져오기
bs_items = ['ifrs-full_Assets', 'ifrs-full_CurrentAssets', 'ifrs-full_CashAndCashEquivalents',
            'dart_ShortTermTradeReceivable', 'ifrs-full_Inventories', 'ifrs-full_PropertyPlantAndEquipment',
            'ifrs-full_ShorttermBorrowings', 'ifrs-full_IssuedCapital', 'ifrs-full_RetainedEarnings',
            'ifrs-full_IntangibleAssetsOtherThanGoodwill', 'ifrs-full_CurrentLiabilities',
            'ifrs-full_Equity', 'ifrs-full_Liabilities']
get_fs_data(wb['Data_bs'], bs_items, df_list)

# # 손익계산서 Data 가져오기
is_items = ['ifrs-full_Revenue', 'ifrs-full_GrossProfit', 'dart_OperatingIncomeLoss', 'ifrs-full_ProfitLoss']
get_fs_data(wb['Data_is'], is_items, df_list)

# 현금흐름 Data 가져오기
cf_items = ['ifrs-full_CashFlowsFromUsedInOperatingActivities',
            'ifrs-full_InterestPaidClassifiedAsOperatingActivities',
            'ifrs-full_CashFlowsFromUsedInInvestingActivities',
            'ifrs-full_CashFlowsFromUsedInFinancingActivities',
            'ifrs-full_ProceedsFromSalesOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities',
            'ifrs-full_PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities']
get_fs_data(wb['Data_cf'], cf_items, df_list)

# 각 데이터프레임 합치기
total_df = pd.concat(df_list, axis=1)

# 합쳐진 데이터 프레임 행/열 바꾸기
total_df = total_df.transpose()

# 재무비율 계산하여 행 추가하기
total_df.loc['Free Cash Flow'] = total_df.loc['ifrs-full_CashFlowsFromUsedInOperatingActivities'] + \
                                 total_df.loc['ifrs-full_ProceedsFromSalesOfPropertyPlantAndEquipmentClassified' \
                                              'AsInvestingActivities'] - \
                                 total_df.loc['ifrs-full_PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvesting' \
                                              'Activities']
total_df.loc['재고자산회전율'] = total_df.loc['ifrs-full_Revenue'] / total_df.loc['ifrs-full_Inventories']
total_df.loc['이자보상배율'] = total_df.loc['dart_OperatingIncomeLoss'] / total_df.loc['ifrs-full_Interest' \
                                                                                 'PaidClassifiedAsOperatingActivities']
total_df.loc['ROA'] = total_df.loc['ifrs-full_ProfitLoss'] / total_df.loc['ifrs-full_Assets']
total_df.loc['ROE'] = total_df.loc['ifrs-full_ProfitLoss'] / total_df.loc['ifrs-full_Equity']
total_df.loc['유동비율'] = total_df.loc['ifrs-full_CurrentLiabilities'] / total_df.loc['ifrs-full_CurrentAssets']
total_df.loc['매출총이익율'] = total_df.loc['ifrs-full_GrossProfit'] / total_df.loc['ifrs-full_Revenue']
total_df.loc['영업이익율'] = total_df.loc['dart_OperatingIncomeLoss'] / total_df.loc['ifrs-full_Revenue']
total_df.loc['순이익율'] = total_df.loc['ifrs-full_ProfitLoss'] / total_df.loc['ifrs-full_Revenue']

# 컬럼과 항목 List의 길이차이
dif_len = len(total_df.iloc[:, [1]]) - len(item_list)

for dif in range(dif_len):
    item_list.append(0)

total_df['항목'] = item_list
col1 = total_df.columns[0:len(total_df.columns) - 1].to_list()
col2 = total_df.columns[-1:].to_list()
total_df = total_df[col2 + col1]

print(total_df)
total_df.to_excel('fs_analy_h.xlsx')

wb_fs = load_workbook('fs_analy_h.xlsx')
ws_fs = wb_fs.worksheets[0]

print(ws_fs)

# 테두리를 지정하기 위해 변수 지정
box = Border(left=Side(border_style='thin', color='FF000000'),
             right=Side(border_style='thin', color='FF000000'),
             top=Side(border_style='thin', color='FF000000'),
             bottom=Side(border_style='thin', color='FF000000'))

# 컬럼 너비를 조정
col_fs = 1
while ws_fs.cell(2, col_fs).value != None:
    ws_fs.column_dimensions[get_column_letter(col_fs)].width = 20
    col_fs += 1

    # 숫자를 천단위 구분 표시
    row_fs = 2
    while row_fs < ws_fs.max_row + 1:
        ws_fs.cell(row_fs, col_fs).number_format = '#,##0'
        ws_fs.cell(row_fs, col_fs).border = box
        row_fs += 1

    # 숫자를 백분위로 표시
    row_fp = 28
    while row_fp < ws_fs.max_row + 1:
        ws_fs.cell(row_fp, col_fs).number_format = '0.00%'
        row_fp += 1

for row in range(1, ws_fs.max_row + 1):
    if ws_fs.cell(row, 2).value == 0:
        ws_fs.cell(row, 2).value = ws_fs.cell(row, 1).value

ws_fs.cell(1, 1).value = file
wb_fs.save('fs_analy_h.xlsx')
wb_fs.close()