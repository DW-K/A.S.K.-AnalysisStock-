import json
import Path
import os
from openpyxl import load_workbook

import subProcess_stock

infoPath = Path.RESOURCE_PATH_STOCK_INFO

jsonCrawlFilePath = Path.RESOURCE_PATH_CRAWLING_KEYWORD
jsonCrawlFileName = Path.RESOURCE_FILE_CRAWLING_KEYWORD
jsonCrawl = Path.RESOURCE_CRAWLING_KEYWORD

jsonStockFilePath = Path.RESOURCE_PATH_STOCK_CODE
jsonStockFileName = Path.RESOURCE_FILE_STOCK_CODE
jsonStock = Path.RESOURCE_STOCK_CODE


def makeCrawlJson():
    infoFileList = os.listdir(infoPath)
    infoFileList = [file for file in infoFileList if file.endswith("xlsx")]

    jsonDict = {}
    for file in infoFileList:   # excel file list
        load_wb = load_workbook(fr'{infoPath}\{file}', data_only=True)
        load_ws = load_wb['회사목록']

        company_list = [row[0].value for row in load_ws.rows]   # first cell value of not null cell

        if '회사목록' in company_list:
            company_list.remove('회사목록')

        categoryName = file.split(".")[0]
        jsonDict[categoryName] = {}
        for sheetName in company_list:
            crawlList = []
            load_ws = load_wb[sheetName]

            for row in load_ws.rows:
                crawlList.append(row[0].value)
            jsonDict[categoryName][sheetName] = crawlList

    writeJson(jsonDict, jsonCrawlFilePath, jsonCrawlFileName)


def makeStockJson():
    infoFileList = os.listdir(infoPath)
    infoFileList = [file for file in infoFileList if file.endswith("xlsx")]

    stockDict = {}
    for file in infoFileList:   # excel file list
        load_wb = load_workbook(fr'{infoPath}\{file}', data_only=True)
        load_ws = load_wb['회사목록']

        company_list = [row[0].value for row in load_ws.rows]   # first cell value of not null cell

        if '회사목록' in company_list:
            company_list.remove('회사목록')

        categoryName = file.split(".")[0]

        stockDict[categoryName] = [companyName for companyName in company_list]
        # stockDict[categoryName] = {}
        #
        # for companyName in company_list:
        #     stockDict[categoryName][companyName] = 0

    writeJson(stockDict, jsonStockFilePath, jsonStockFileName)

    subProcess_stock.getStockCode_temp(jsonStockFilePath, jsonStockFileName)

    # subProcess_stock.getStockCode(json.dumps(stockDict, ensure_ascii=False), jsonStockFilePath, jsonStockFileName)


def writeJson(jsonDict, filePath, fileName):
    if os.path.isfile(fr'{filePath}\{fileName}'):
        with open(fr'{filePath}\{fileName}', "w", encoding='UTF8') as json_file:
            json.dump(jsonDict, json_file, ensure_ascii=False)
    else:
        if createJsonFile(filePath, fileName):
            writeJson(jsonDict, filePath, fileName)
        else:
            print("create json file error")


def createJsonFile(filePath, fileName):
    Path.createFolder(filePath)
    with open(fr'{filePath}\{fileName}', "w"):
        pass
    return True


def readJson(fileType=None, filePath=None, fileName=None):
    if fileType == 'stock':
        with open(jsonStock, 'r', encoding='UTF8') as f:
            json_data = json.load(f)
    elif fileType == 'crawl':
        with open(jsonCrawl, 'r', encoding='UTF8') as f:
            json_data = json.load(f)
    elif filePath is not None and fileName is not None:
        with open(fr'{filePath}\{fileName}', 'r', encoding='UTF8') as f:
            json_data = json.load(f)
    else:
        print('wrong params')
        return False

    return json_data


if __name__ == "__main__":
    # crawlDict = readJson('crawl')

    makeStockJson()
